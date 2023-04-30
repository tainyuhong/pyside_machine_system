import sys
import pathlib
import openpyxl
from openpyxl.styles import Border, Side
from PySide6 import QtWidgets, QtCore
from db.db_orm import *
from peewee import fn
from ui.export_excel import Ui_export_form


# 获取机房名称
def get_room():
    room_model = MachineRoom.select(MachineRoom.room_id, MachineRoom.room_name).order_by(
        MachineRoom.room_id).tuples().execute()
    room_data = [i for i in room_model]  # 转换机字典
    # print('机房信息：', room_data)
    return room_data


class ExportExcel(Ui_export_form, QtWidgets.QWidget):
    # 定义字段名字典
    field_dict = {'machine_id': '设备ID', 'machine_name': '设备名称', 'machine_sort_name': '分类名称',
                  'machine_sn': '序列号', 'machine_factory': '设备厂商', 'model': '型号',
                  'machine_roomid': '机房', 'cabinet_name': '机柜编号', 'start_position': '开始U位',
                  'end_position': '结束U位', 'factory_date': '出产日期', 'end_ma_date': '到保日期',
                  'work_are': '业务类型 ', 'run_state': '运行状态 ', 'machine_admin': '管理员',
                  'app_admin': '应用管理员',
                  'mg_ip': '管理IP地址', 'app_ip1': '业务IP', 'bmc_ip': 'BMC IP',
                  'install_date': '上架安装时间',
                  'uninstall_date': '下架时间', 'single_power': '单电源', 'asset_id': '资产编号','system_name': '系统名称',
                  'comments': '备注'}

    def __init__(self, parent=None):
        super(ExportExcel, self).__init__(parent)
        self.setupUi(self)  # 展示报告窗口页
        self.get_field()  # 显示数据库中字段
        # self.room = get_room()     # 获取机房id与名字数据
        self.btn_export.clicked.connect(self.export_exl)  # 导出按钮事件

    # 获取要导出的字段
    def get_field(self):
        # 将多选框添加到列表控件中
        for i in self.field_dict.values():
            box = QtWidgets.QCheckBox(i)  # 实例化一个QCheckBox，把文字传进去
            box.setChecked(True)  # 默认设置为钩选状态
            item = QtWidgets.QListWidgetItem()  # 实例化一个Item，QListWidget，不能直接加入QCheckBox
            self.listWidget.addItem(item)  # 把QListWidgetItem加入QListWidget
            self.listWidget.setItemWidget(item, box)  # 再把QCheckBox加入QListWidgetItem

    # 导出按钮事件
    def export_exl(self):
        # print('导出至excel')
        exp_sql = """SELECT {} FROM machine_infos """
        new_data = dict(zip(self.field_dict.values(), self.field_dict.keys()))  # 将字典key/value进行反转
        count = self.listWidget.count()     # 列数
        chooses_list = []  # 选择的字段内容
        choose_name = []  # 选择
        # 动态生成机房名与ID转换sql
        room_sql = 'case'
        for i in get_room():
            room_sql = room_sql + "  when machine_roomid='{}' then '{}' ".format(i[0], i[1])
        room_sql = room_sql + ' end as roomid'
        # print('room_sql', room_sql)

        for i in range(count):
            # print('列表项：',self.listWidget.item(i))
            chk_item = self.listWidget.itemWidget(self.listWidget.item(i))  # 获取复选框对象
            # print('列表项：',chk_item.isChecked())
            # 判断复选框是否选择
            if chk_item.isChecked():
                chooses_list.append(chk_item.text())
                # print('数据库字段名：',new_data[chk_item.text()])

                # 设置业务类型字段,将状态码转换为字符显示
                if new_data[chk_item.text()] == 'work_are':
                    new_data[chk_item.text()] = "case when work_are = '1' then '生产' when work_are ='2' then '电渠' \
                    when work_are ='3' then '灾备' when work_are ='4' then '开发' when work_are ='5' then '备份' \
                     when work_are ='6' then '分行' end as workare"
                # 设置运行状态字段,将状态码转换为字符显示
                elif new_data[chk_item.text()] == 'run_state':
                    new_data[chk_item.text()] = "case when run_state = '1' then '运行' when run_state ='2' then '断网'\
                                 when run_state ='3' then '关机' when run_state ='4' then '下架' \
                                 when run_state ='5' then '未加电' end as runstat"
                # 机房ID与机房名称的转换
                elif new_data[chk_item.text()] == 'machine_roomid':
                    new_data[chk_item.text()] = room_sql
                # 设备是否单电源字段，将状态码转换为字符显示
                elif new_data[chk_item.text()] == 'single_power':
                    new_data[chk_item.text()] = "case when single_power='1'  then '是' \
                    when single_power='0' then '否' end as is_single_power "

                choose_name.append(new_data[chk_item.text()])
        # print('选择的字段值内容：', chooses_list)
        # print('数据库字段名：', choose_name)

        # # 查询数据
        # export_data_model = MachineInfos.select(*choose_name)
        # print('查询到的数据：',export_data_model)

        # 拼接sql
        if len(choose_name) > 0:
            select_sql = exp_sql.format(', '.join(choose_name))  # 通过join链接列表中各字段值
            # print('查询语句：',select_sql)
            # 执行查询，获取所有数据
            try:
                data = database.execute_sql(select_sql).fetchall()
            except Exception as e:
                logging.critical('获取数据库错误：', e)
            else:
                # print('查询到的数据：',data)
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = '设备信息'
                ws.append(chooses_list)  # 写入表格标题栏
                # 创建边框线对象
                border = Border(left=Side(style='thin',color='000000'),
                                right=Side(style='thin',color='000000'),
                                top=Side(style='thin',color='000000'),
                                bottom=Side(style='thin',color='000000'))

                # 向表格中写入数据
                for row in data:
                    ws.append(row)  # 把查询到的数据写入到表中

                # 批量对单元格进行格式化
                for r in ws.iter_rows(min_row=1,max_col=ws.max_column,max_row=ws.max_row):
                    for c in r:
                        c.border=border     # 设置单元格边框样式

                # 保存
                filepath, filetype = QtWidgets.QFileDialog.getSaveFileUrl(self, '保存导出文件', filter='.xlsx')
                # print('保存文件路径：',filepath,filepath.toLocalFile())
                if filepath!='':
                    if pathlib.Path(filepath.toLocalFile()).suffix == '.xlsx':
                        # 将文件保存到指定目录
                        try:
                            wb.save(filepath.toLocalFile())
                        except Exception as e:
                            QtWidgets.QMessageBox.critical(self, '保存文件', '保存文件错误！{}'.format(e))
                        else:
                            wb.close()
                            QtWidgets.QMessageBox.information(self, '保存文件', '保存文件成功！！')
                    else:
                        # 如果用户没有在文件名后加后缀名，则系统自动加上
                        save_file = filepath.toLocalFile() + '.xlsx'
                        # 将文件从系统下载到指定目录
                        try:
                            wb.save(save_file)
                        except Exception as e:
                            QtWidgets.QMessageBox.critical(self, '保存文件', '保存文件错误！{}'.format(e))
                        else:
                            wb.close()
                            QtWidgets.QMessageBox.information(self, '保存文件', '保存文件成功！！')
                else:
                    print('取消导出！')
        else:
            QtWidgets.QMessageBox.warning(self, '选择需要导出的字段', '未选择需要导出的字段，请选择！！')


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    report_win = ExportExcel()
    report_win.show()
    sys.exit(app.exec())
