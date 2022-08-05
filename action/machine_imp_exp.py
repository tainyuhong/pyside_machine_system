from pathlib import Path
import sys
import openpyxl
from PySide6 import QtWidgets, QtGui, QtCore
from ui.machine_import import *
from db.db_orm import *


class UiImport(QtWidgets.QWidget, Ui_import_machine):
    def __init__(self, parent=None):
        super(UiImport, self).__init__(parent)
        self.setupUi(self)
        self.select_btn.clicked.connect(self.open_file)
        self.import_btn.clicked.connect(lambda: self.imp_file(self.path_le.text()))
        self.bt_download.clicked.connect(self.download_template)

    def open_file(self):
        # 创建文件选择框实例，并接收文件路径信息
        filename, filetype = QtWidgets.QFileDialog.getOpenFileName(self, '选择文件', str(Path.cwd()),'excel文件 (*.xlsx)')
        # print(filename)
        self.path_le.setText(filename)  # 将获取的选择的文件路径显示至文本框

    def imp_file(self, path):
        # 导入excel文件

        if path != '':
            _ = QtWidgets.QMessageBox.question(self, '设备信息导入', '是否需要导入设备信息！')
            if _ == QtWidgets.QMessageBox.Yes:
                # print('路径', path)
                try:
                    wb = openpyxl.load_workbook(r'{}'.format(path))
                    # print('path', path)
                    sh = wb.active
                except Exception as e:
                    QtWidgets.QMessageBox.critical(self, '导入失败', '错误：{}'.format(e))
                else:
                    rowdata = []
                    data = []
                    for row in range(3, sh.max_row + 1):
                        for col in range(1, sh.max_column + 1):
                            rowdata.append(sh.cell(row, col).value)  # 将单元格数据添加到行列表中
                        # print(rowdata)
                        data.append(rowdata)
                        rowdata = []  # 置空列表添加新的行
                    # print(data)
                    try:
                        with database.atomic():
                            MachineInfos.insert_many(data, fields=['machine_id', 'machine_name', 'machine_sort_name',
                                                                   'machine_sn', 'machine_factory', 'model',
                                                                   'machine_roomid', 'cabinet_name',
                                                                   'start_position', 'end_position', 'factory_date',
                                                                   'end_ma_date', 'work_are', 'run_state',
                                                                   'machine_admin', 'app_admin',
                                                                   'mg_ip', 'app_ip1', 'bmc_ip', 'install_date',
                                                                   'uninstall_date', 'single_power',
                                                                   'comments']).execute()
                    except Exception as e:
                        print('插入数据库中错误：', e)
                        QtWidgets.QMessageBox.warning(self, '设备信息导入', '导入失败，未导入任何信息！')
                    else:

                        # print('导入成功！！')
                        QtWidgets.QMessageBox.information(self, '设备信息导入','{}条记录插入成功 '.format(len(data)))

            else:
                print('取消导入')
        else:
            QtWidgets.QMessageBox.warning(self, '设备信息导入', '请选择需要导入的文件！')

    def download_template(self):
        # print('点击了模板下载')
        # print(Path(__file__))
        template_path = Path(__file__).parents[1]/'machine_template'/'machine_infos.xlsx'
        # print(template_path)
        file,filetype = QtWidgets.QFileDialog.getSaveFileUrl(self,'下载设备导入模板',str(template_path),'.xlsx')
        # print('保存的文件名：',file,file.toLocalFile())
        save_file = file.toLocalFile()
        # print('文件后缀名：',Path(save_file).suffix)
        if Path(save_file).suffix == '.xlsx':
            # 将文件从系统下载到指定目录
            try:
                Path(save_file).write_bytes(template_path.read_bytes())
            except Exception as e:
                QtWidgets.QMessageBox.critical(self,'下载模板','下载模板错误！{}'.format(e))
            else:
                QtWidgets.QMessageBox.information(self, '下载模板', '下载模板成功！！')
        else:
            # 如果用户没有在文件名后加后缀名，则系统自动加上
            save_file = save_file+'.xlsx'
            # 将文件从系统下载到指定目录
            try:
                Path(save_file).write_bytes(template_path.read_bytes())
            except Exception as e:
                QtWidgets.QMessageBox.critical(self, '下载模板', '下载模板错误！{}'.format(e))
            else:
                QtWidgets.QMessageBox.information(self, '下载模板', '下载模板成功！！')


if __name__ == '__main__':
    app = QtWidgets.QApplication(sys.argv)
    import_win = UiImport()
    import_win.show()
    sys.exit(app.exec())
